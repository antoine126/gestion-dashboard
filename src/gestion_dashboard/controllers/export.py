"""Export utilities: Excel (.xlsx) and JSON."""
from __future__ import annotations

import io
from typing import List

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from gestion_dashboard.models.budget import (
    AnalyseBudget,
    CategorieDepense,
    Depense,
    SoldeMensuel,
)
from gestion_dashboard.models.enums import MOIS

_HEADER_FILL = PatternFill(start_color="5B4FBE", end_color="5B4FBE", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_TITLE_FONT = Font(size=13, bold=True, color="2C3E50")
_ALT_FILL = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")


def _auto_col_width(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 45)


def export_synthese_mensuelle_xlsx(
    solde: SoldeMensuel,
    analyses: List[AnalyseBudget],
    categories: List[CategorieDepense],
    depenses: List[Depense],
) -> bytes:
    """
    Generate an Excel workbook for a monthly summary.

    Parameters
    ----------
    solde : SoldeMensuel
        Monthly balance summary.
    analyses : List[AnalyseBudget]
        Budget vs actual per category.
    categories : List[CategorieDepense]
        All expense categories.
    depenses : List[Depense]
        Expense journal for the month.

    Returns
    -------
    bytes
        Excel file content as bytes, ready for ``st.download_button``.
    """
    wb = openpyxl.Workbook()

    # ── Sheet 1: Monthly summary ───────────────────────────────────────────
    ws = wb.active
    ws.title = f"Synthèse {MOIS[solde.mois - 1]}"
    ws["A1"] = f"Budget Personnel — {MOIS[solde.mois - 1]} {solde.annee}"
    ws["A1"].font = _TITLE_FONT
    ws.append([])

    ws.append(["RÉCAPITULATIF MENSUEL"])
    ws[ws.max_row][0].font = Font(bold=True, color="5B4FBE")
    ws.append(["Revenus totaux",       f"{solde.total_revenus:.2f} €"])
    ws.append(["Charges fixes",        f"-{solde.total_charges_fixes:.2f} €"])
    ws.append(["Épargne programmée",   f"-{solde.total_epargne:.2f} €"])
    ws.append(["Solde disponible",     f"{solde.solde_disponible:.2f} €"])
    ws.append([])

    headers = ["Catégorie", "Budgété (€)", "Réel (€)", "Écart (€)", "% Consommé"]
    ws.append(headers)
    for cell in ws[ws.max_row]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for i, a in enumerate(analyses):
        row = [
            f"{a.categorie_icone} {a.categorie_nom}",
            round(a.budgete, 2),
            round(a.reel, 2),
            round(a.ecart, 2),
            f"{a.pourcentage_consomme:.1f}%",
        ]
        ws.append(row)
        if i % 2 == 0:
            for cell in ws[ws.max_row]:
                cell.fill = _ALT_FILL

    _auto_col_width(ws)

    # ── Sheet 2: Expense journal ───────────────────────────────────────────
    ws2 = wb.create_sheet("Journal des dépenses")
    cat_map = {c.id: f"{c.icone} {c.nom}" for c in categories}
    headers2 = ["Date", "Libellé", "Catégorie", "Montant (€)", "Mode de paiement", "Note"]
    ws2.append(headers2)
    for cell in ws2[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT

    for i, d in enumerate(sorted(depenses, key=lambda x: x.date)):
        ws2.append([
            d.date, d.libelle,
            cat_map.get(d.categorie_id, "Inconnu"),
            round(d.montant, 2), d.mode_paiement, d.note,
        ])
        if i % 2 == 0:
            for cell in ws2[ws2.max_row]:
                cell.fill = _ALT_FILL

    _auto_col_width(ws2)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
